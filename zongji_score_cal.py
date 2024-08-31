import openpyxl
from datetime import datetime

filename = "综基历年真题做题记录.xlsx"    #excel文件名
end_sent = "加油"                  #单卷记录结束鼓励自己的话

def get_deduct_points(a):
    """根据错题号返回扣分点数"""
    if 1 <= a <= 30:
        return 1
    elif 31 <= a <= 55:
        return 0.8
    elif 56 <= a <= 65:
        return 1
    elif 66 <= a <= 70:
        return 0.9
    elif 71 <= a <= 75:
        return 0.7
    elif 76 <= a <= 80:
        return 0.8
    elif 81 <= a <= 85:
        return 0.9
    elif 86 <= a <= 95:
        return 1
    elif 96 <= a <= 100:
        return 0.9
    elif 101 <= a <= 110:
        return 1
    else:
        return None

def record_results(year, month, total_points, entered_answers, correct_answers):
    """将结果记录到 Excel 文件"""
    
    
    # 打开或创建 Excel 文件
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 选择活动工作表或创建新工作表
    if '记录' in wb.sheetnames:
        sheet = wb['记录']
    else:
        sheet = wb.active
        sheet.title = '记录'

    # 查找下一空列
    col = sheet.max_column + 1

    # 记录标题行
    sheet.cell(row=1, column=col, value="卷子")
    sheet.cell(row=2, column=col, value="做题日期")
    sheet.cell(row=3, column=col, value="总分")
    sheet.cell(row=4, column=col, value="题号")
    sheet.cell(row=4, column=col + 1, value="错选项")
    sheet.cell(row=4, column=col + 2, value="正确选项")
    
    # 记录卷子信息
    sheet.cell(row=1, column=col + 1, value=f"{year}年{month}月")
    # 记录日期
    sheet.cell(row=2, column=col + 1, value=datetime.now().strftime("%Y-%m-%d"))
    # 记录总分
    sheet.cell(row=3, column=col + 1, value=f"{total_points:.1f}")

    # 记录错题及答案
    for i in range(1, 111):
        sheet.cell(row=i + 4, column=col, value=i)
        answer = entered_answers.get(i, '')
        correct_answer = correct_answers.get(i, '')
        if answer:
            sheet.cell(row=i + 4, column=col + 1, value=f"错选：{answer}")
        else:
            sheet.cell(row=i + 4, column=col + 1, value="")
        if correct_answer:
            sheet.cell(row=i + 4, column=col + 2, value=f"正确：{correct_answer}")
        else:
            sheet.cell(row=i + 4, column=col + 2, value="")
        # 添加空列以分隔数据
        sheet.cell(row=i + 4, column=col + 3, value="")

    # 保存文件
    wb.save(filename)

def main():
    while True:
        print("--------------------------")
        total_points = 100.0
        last_deduct_points = 0
        entered_answers = {}  # 用于存储错题号和对应的错选项
        correct_answers = {}  # 用于存储每题的正确选项

        # 输入年份和月份
        year = input("请输入真题的年份: ")
        month = input("请输入真题的月份: ")

        while True:
            print("输入错题号(1-110)    '0'_上次输错不算  'end'_输入完毕")
            user_input = input("->")

            if user_input == 'end':
                break

            try:
                mistake_number = int(user_input)
            except ValueError:
                print("输入无效，输入(1-110)或'end'结束")
                continue

            if mistake_number < 0 or mistake_number > 110:
                print("输错了，别逗")
                continue

            if mistake_number == 0:
                print("行，上题不算")
                total_points += last_deduct_points
                # 从已输入错题号中移除最后一次错题
                if entered_answers:
                    last_mistake = max(entered_answers.keys())
                    entered_answers.pop(last_mistake, None)
                    correct_answers.pop(last_mistake, None)
            elif mistake_number in entered_answers:
                print("输过这题了，别逗")
            else:
                deduct_points = get_deduct_points(mistake_number)
                if deduct_points is None:
                    print("输错了，别逗")
                    continue
                
                # 输入错选项
                while True:
                    print("输入错选项（A, B, C, D）:")
                    answer = input("->").upper()
                    if answer in {'A', 'B', 'C', 'D'}:
                        break
                    else:
                        print("错选项无效，请重新输入（A, B, C, D）")
                
                # 输入正确选项
                while True:
                    print("输入正确选项（A, B, C, D）:")
                    correct_answer = input("->").upper()
                    if correct_answer in {'A', 'B', 'C', 'D'}:
                        break
                    else:
                        print("正确选项无效，请重新输入（A, B, C, D）")
                
                total_points -= deduct_points
                last_deduct_points = deduct_points
                entered_answers[mistake_number] = answer
                correct_answers[mistake_number] = correct_answer

            print(f"当前总分: {total_points:.1f}")
            print()

        # 记录结果到 Excel 文件
        record_results(year, month, total_points, entered_answers, correct_answers)

        print(f"最后总分: {total_points:.1f}  , {end_sent}")
        print()

if __name__ == "__main__":
    main()
